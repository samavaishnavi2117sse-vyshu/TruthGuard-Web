# generate_dummy_report.py
"""Generate a dummy test report with 100 passed test cases.
Creates an Excel file and a markdown summary in the project root.
"""
import os, datetime, openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

NUM_TESTS = 100

def create_excel(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    headers = ["#", "Test ID", "Module", "Test Name", "Status", "Duration (ms)"]
    ws.append(headers)
    for i in range(1, NUM_TESTS + 1):
        ws.append([i, f"TC-{i:03d}", "DummyModule", f"Dummy test {i}", "PASSED", 0])
    # Simple styling
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    wb.save(path)
    print(f"Excel report saved to {path}")

def create_markdown(path):
    md = "# 📊 Dummy Test Report\n\n"
    md += f"Generated on {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
    md += "| # | Test ID | Module | Test Name | Status |\n"
    md += "|---|---------|--------|-----------|--------|\n"
    for i in range(1, NUM_TESTS + 1):
        md += f"| {i} | TC-{i:03d} | DummyModule | Dummy test {i} | PASSED |\n"
    with open(path, "w", encoding="utf-8") as f:
        f.write(md)
    print(f"Markdown summary saved to {path}")

if __name__ == "__main__":
    base_dir = os.path.dirname(os.path.abspath(__file__))
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(base_dir, f"Dummy_Test_Report_{ts}.xlsx")
    md_path = os.path.join(base_dir, "dummy_test_summary.md")
    create_excel(excel_path)
    create_markdown(md_path)
