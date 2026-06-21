import os
import sys
import time
import subprocess
from datetime import datetime

# ── UTF-8 output (emoji support on Windows) ───────────────────────────────────
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')

# ── Package imports ───────────────────────────────────────────────────────────
try:
    from appium import webdriver
    from appium.options.android import UiAutomator2Options
    from appium.webdriver.common.appiumby import AppiumBy
except ImportError:
    print("Required packages missing. Run: pip install -r requirements.txt")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("openpyxl missing. Run: pip install openpyxl")
    sys.exit(1)

# ═══════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════
IS_LINUX = sys.platform.startswith('linux')
IS_CI    = os.environ.get('CI', '').lower() in ('true', '1')

SDK_PATH = os.environ.get(
    "ANDROID_HOME",
    "/usr/local/lib/android/sdk" if IS_LINUX
    else r"C:\Users\HP\AppData\Local\Android\Sdk"
)

ADB_PATH      = "adb" if IS_LINUX else os.path.join(SDK_PATH, "platform-tools", "adb.exe")
EMULATOR_PATH = ""    if IS_LINUX else os.path.join(SDK_PATH, "emulator", "emulator.exe")

_apk_env = os.environ.get("APK_PATH", "")
if _apk_env:
    APK_PATH = _apk_env
elif IS_LINUX:
    APK_PATH = os.path.join(
        os.environ.get("GITHUB_WORKSPACE", ""),
        "app/build/outputs/apk/debug/app-debug.apk"
    )
else:
    APK_PATH = r"C:\Users\HP\Projects\TRUTH GUARD\app\build\outputs\apk\debug\app-debug.apk"

AVD_NAME     = os.environ.get("AVD_NAME", "Pixel_6")
APP_PACKAGE  = "com.samavaishnavi.truthguard"
APPIUM_PORT  = 4723
APPIUM_HOST  = "127.0.0.1"

# ═══════════════════════════════════════════════════════════════════════════════
#  REPORT COLOUR PALETTE
# ═══════════════════════════════════════════════════════════════════════════════
COLORS = {
    'headerBg':    '1A1A2E',
    'headerText':  'FFFFFF',
    'subHeaderBg': '16213E',
    'moduleBg':    '0F3460',
    'moduleText':  'E94560',
    'passedBg':    'D4EDDA',
    'passedText':  '155724',
    'failedBg':    'F8D7DA',
    'failedText':  '721C24',
    'rowAlt':      'F8F9FA',
    'rowNorm':     'FFFFFF',
    'accent':      '0F3460',
    'statPass':    '28A745',
    'statFail':    'DC3545',
    'statTotal':   '17A2B8',
    'border':      'BDC3C7',
}

# ═══════════════════════════════════════════════════════════════════════════════
#  RESULTS STORE
# ═══════════════════════════════════════════════════════════════════════════════
results     = []
suite_start = time.time()

# ═══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def log(status, test_id, name, duration, err=''):
    icon = '✅' if status == 'PASSED' else '❌'
    snippet = f" → {err[:90]}" if err else ""
    print(f"  {icon} [{test_id}] {name} ({duration}ms){snippet}")


def record(module, test_id, name, desc, status, duration, error=''):
    results.append({
        'module': module, 'id': test_id, 'name': name,
        'desc': desc, 'status': status, 'duration': duration, 'error': error
    })
    log(status, test_id, name, duration, error)


def tc(module, test_id, name, desc, fn):
    """Run one test case and record the result."""
    t0 = time.time()
    try:
        fn()
        record(module, test_id, name, desc, 'PASSED', int((time.time() - t0) * 1000))
    except Exception as e:
        record(module, test_id, name, desc, 'FAILED', int((time.time() - t0) * 1000), str(e))

# ═══════════════════════════════════════════════════════════════════════════════
#  ELEMENT FINDERS / INTERACTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def wait_for_element(driver, by, selector, timeout=8):
    """Wait for an element to be present using the given strategy."""
    end = time.time() + timeout
    while time.time() < end:
        try:
            return driver.find_element(by, selector)
        except Exception:
            time.sleep(0.3)
    raise AssertionError(f"Element with {by}='{selector}' not found after {timeout}s")


def find_text(driver, text, timeout=8):
    """Find an element containing given text using multiple UIAutomator strategies."""
    end = time.time() + timeout
    while time.time() < end:
        # Strategy 1: UIAutomator textContains
        try:
            return driver.find_element(AppiumBy.ANDROID_UIAUTOMATOR,
                                       f'new UiSelector().textContains("{text}")')
        except Exception:
            pass
        # Strategy 2: XPath contains
        try:
            return driver.find_element(AppiumBy.XPATH,
                                       f'//*[contains(@text, "{text}")]')
        except Exception:
            pass
        time.sleep(0.3)
    raise AssertionError(f"Text '{text}' not found on screen after {timeout}s")



def assert_visible(driver, text, timeout=8):
    find_text(driver, text, timeout)


def assert_absent(driver, text, timeout=3):
    """Assert that text does NOT appear on screen."""
    try:
        find_text(driver, text, timeout=timeout)
        raise AssertionError(f"Text '{text}' should NOT be visible but was found")
    except AssertionError as e:
        if "should NOT" in str(e):
            raise
        # text not found = correct
        pass


def click_btn(driver, text, timeout=8):
    el = find_text(driver, text, timeout)
    el.click()
    time.sleep(0.6)


def enter_text(driver, text):
    """Clear and optionally type into the single EditText on screen."""
    # Try locating the EditText using multiple strategies for better reliability
    try:
        el = wait_for_element(driver, AppiumBy.CLASS_NAME, "android.widget.EditText")
    except Exception:
        # Fallback to UIAutomator selector based on hint or resource-id if class name fails
        el = wait_for_element(
            driver,
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().className("android.widget.EditText")'
        )
    el.clear()
    if text:
        el.send_keys(text)
    time.sleep(0.3)



def go_back(driver, pause=0.5):
    driver.back()
    time.sleep(pause)


def analyze(driver, text):
    """Enter text, click Analyze, and wait for result card to appear."""
    enter_text(driver, text)
    click_btn(driver, "Analyze")
    # Wait for result card text indicating analysis completed
    wait_for_element(driver, AppiumBy.ANDROID_UIAUTOMATOR, f'new UiSelector().textContains("Analysis Result")', timeout=15)
    time.sleep(0.6)


def swipe_down(driver):
    # Use native driver.swipe for swipe down
    size = driver.get_window_size()
    start_y = int(size['height'] * 0.8)
    end_y = int(size['height'] * 0.2)
    x = int(size['width'] / 2)
    driver.swipe(x, start_y, x, end_y, 500)


def swipe_up(driver):
    # Use native driver.swipe for swipe up
    size = driver.get_window_size()
    start_y = int(size['height'] * 0.2)
    end_y = int(size['height'] * 0.8)
    x = int(size['width'] / 2)
    driver.swipe(x, start_y, x, end_y, 500)

# ═══════════════════════════════════════════════════════════════════════════════
#  ENVIRONMENT SETUP
# ═══════════════════════════════════════════════════════════════════════════════
def start_emulator():
    print("\n🔍 Checking for running emulator/device...")
    res = subprocess.run([ADB_PATH, "devices"], capture_output=True, text=True)
    lines = [l.strip() for l in res.stdout.splitlines()
             if l.strip() and "List of devices" not in l]
    if any("device" in l or "emulator" in l for l in lines):
        print("🟢 Emulator/device already running.")
        return

    if IS_CI:
        print("⏳ CI mode: waiting for emulator from runner action...")
        subprocess.run([ADB_PATH, "wait-for-device"], timeout=180)
        for _ in range(60):
            r = subprocess.run([ADB_PATH, "shell", "getprop", "sys.boot_completed"],
                               capture_output=True, text=True)
            if "1" in r.stdout:
                print("🟢 Emulator ready.")
                time.sleep(3)
                return
            time.sleep(3)
        print("⚠️  Boot-check timed out — proceeding.")
        return

    if not EMULATOR_PATH:
        raise RuntimeError("EMULATOR_PATH not configured.")
    print(f"🚀 Launching emulator: {AVD_NAME}")
    subprocess.Popen([EMULATOR_PATH, "-avd", AVD_NAME, "-delay-adb"])
    subprocess.run([ADB_PATH, "wait-for-device"])
    for _ in range(60):
        r = subprocess.run([ADB_PATH, "shell", "getprop", "sys.boot_completed"],
                           capture_output=True, text=True)
        if "1" in r.stdout:
            print("🟢 Emulator booted.")
            time.sleep(3)
            return
        time.sleep(2)
    print("⚠️  Boot-check timed out.")


def start_appium_server():
    if IS_CI:
        print("ℹ️  CI mode: Appium managed by workflow.")
        return None, None
    print("🚀 Starting Appium server...")
    log_f = open("appium_server.log", "w")
    env   = os.environ.copy()
    env["ANDROID_HOME"]     = SDK_PATH
    env["ANDROID_SDK_ROOT"] = SDK_PATH
    env["JAVA_HOME"]        = r"C:\Program Files\Android\Android Studio\jbr"
    proc = subprocess.Popen(
        ["cmd", "/c", "npx", "appium",
         "--port", str(APPIUM_PORT), "--address", APPIUM_HOST],
        env=env, stdout=log_f, stderr=log_f
    )
    time.sleep(8)
    print("🟢 Appium server started.")
    return proc, log_f


# ═══════════════════════════════════════════════════════════════════════════════
#  TEST SUITE — 135 TEST CASES · 8 MODULES
# ═══════════════════════════════════════════════════════════════════════════════
def run_all_tests(driver):

    # ══════════════════════════════════════════════════════════════════════════
    # MODULE 1 · APP LAUNCH & HOME SCREEN  (TC-001 – TC-022)
    # ══════════════════════════════════════════════════════════════════════════
    print("\n📋 MODULE 1: App Launch & Home Screen")

    tc("App Launch", "TC-001", "App launches without crash",
       "MainActivity starts successfully",
       lambda: driver.activate_app(APP_PACKAGE))

    tc("App Launch", "TC-002", "Shield emoji '🛡️' visible on Home",
       "Emoji Text element present",
       lambda: assert_visible(driver, "🛡️"))

    tc("App Launch", "TC-003", "Title 'TRUTHGUARD' visible",
       "Bold title text present",
       lambda: assert_visible(driver, "TRUTHGUARD"))

    tc("App Launch", "TC-004", "Subtitle 'AI Powered Fake News Detection' visible",
       "Subtitle text present",
       lambda: assert_visible(driver, "AI Powered Fake News Detection"))

    tc("App Launch", "TC-005", "'Version 1.0' footer label visible",
       "Version text at bottom of screen",
       lambda: assert_visible(driver, "Version 1.0"))

    tc("App Launch", "TC-006", "'Verify News' button visible on Home",
       "Button with label '🔍 Verify News' present",
       lambda: assert_visible(driver, "Verify News"))

    tc("App Launch", "TC-007", "'Trending News' button visible on Home",
       "Button with label '📰 Trending News' present",
       lambda: assert_visible(driver, "Trending News"))

    tc("App Launch", "TC-008", "'Dashboard' button visible on Home",
       "Button with label '📊 Dashboard' present",
       lambda: assert_visible(driver, "Dashboard"))

    tc("App Launch", "TC-009", "'About' button visible on Home",
       "Button with label 'ℹ About' present",
       lambda: assert_visible(driver, "About"))

    tc("App Launch", "TC-010", "Home screen has at least 4 buttons",
       "android.widget.Button count >= 4",
       lambda: _check_btn_count(driver))

    tc("App Launch", "TC-011", "'Verify News' button is clickable",
       "Clickable attribute = true",
       lambda: _check_clickable(driver, "Verify News"))

    tc("App Launch", "TC-012", "'Trending News' button is clickable",
       "Clickable attribute = true",
       lambda: _check_clickable(driver, "Trending News"))

    tc("App Launch", "TC-013", "'Dashboard' button is clickable",
       "Clickable attribute = true",
       lambda: _check_clickable(driver, "Dashboard"))

    tc("App Launch", "TC-014", "'About' button is clickable",
       "Clickable attribute = true",
       lambda: _check_clickable(driver, "About"))

    tc("App Launch", "TC-015", "Tap 'Verify News' → Verify screen opens",
       "Click button, Verify header appears",
       lambda: (_nav_to(driver, "Verify News", "Verify News")))

    go_back(driver)

    tc("App Launch", "TC-016", "Tap 'Trending News' → Trending screen opens",
       "Click button, '📰 Trending News' header appears",
       lambda: _nav_to(driver, "Trending News", "📰 Trending News"))

    go_back(driver)

    tc("App Launch", "TC-017", "Tap 'Dashboard' → Dashboard screen opens",
       "Click button, '📊 Dashboard' header appears",
       lambda: _nav_to(driver, "Dashboard", "📊 Dashboard"))

    go_back(driver)

    tc("App Launch", "TC-018", "Tap 'About' → About screen opens",
       "Click button, 'TruthGuard' about title appears",
       lambda: _nav_to(driver, "About", "TruthGuard"))

    go_back(driver)

    tc("App Launch", "TC-019", "Back from Verify → Home",
       "TRUTHGUARD title visible after back press",
       lambda: (_nav_to(driver, "Verify News", "Verify News"),
                go_back(driver),
                assert_visible(driver, "TRUTHGUARD")))

    tc("App Launch", "TC-020", "Back from Trending → Home",
       "TRUTHGUARD title visible after back press",
       lambda: (_nav_to(driver, "Trending News", "Trending News"),
                go_back(driver),
                assert_visible(driver, "TRUTHGUARD")))

    tc("App Launch", "TC-021", "Back from Dashboard → Home",
       "TRUTHGUARD title visible after back press",
       lambda: (_nav_to(driver, "Dashboard", "Dashboard"),
                go_back(driver),
                assert_visible(driver, "TRUTHGUARD")))

    tc("App Launch", "TC-022", "Back from About → Home",
       "TRUTHGUARD title visible after back press",
       lambda: (_nav_to(driver, "About", "TruthGuard"),
                go_back(driver),
                assert_visible(driver, "TRUTHGUARD")))

    # ══════════════════════════════════════════════════════════════════════════
    # MODULE 2 · VERIFY NEWS — UI STRUCTURE  (TC-023 – TC-040)
    # ══════════════════════════════════════════════════════════════════════════
    print("\n📋 MODULE 2: Verify News — UI Structure")
    click_btn(driver, "Verify News")

    tc("Verify UI", "TC-023", "Verify screen header 'Verify News' present",
       "Screen title text visible",
       lambda: assert_visible(driver, "Verify News"))

    tc("Verify UI", "TC-024", "'Paste News Here' field label visible",
       "OutlinedTextField label present",
       lambda: assert_visible(driver, "Paste News Here"))

    tc("Verify UI", "TC-025", "EditText input field present",
       "android.widget.EditText element found",
       lambda: driver.find_element(AppiumBy.CLASS_NAME, "android.widget.EditText"))

    tc("Verify UI", "TC-026", "'Analyze' button visible",
       "Analyze button present on Verify screen",
       lambda: assert_visible(driver, "Analyze"))

    tc("Verify UI", "TC-027", "Analyze button is clickable",
       "Clickable attribute = true",
       lambda: _check_clickable_nodestructive(driver, "Analyze"))

    tc("Verify UI", "TC-028", "EditText is enabled",
       "enabled attribute = true",
       lambda: _check_field_enabled(driver))

    tc("Verify UI", "TC-029", "Result card absent before first analysis",
       "'Analysis Result' text not visible initially",
       lambda: assert_absent(driver, "Analysis Result", timeout=3))

    tc("Verify UI", "TC-030", "User can type text into EditText",
       "Text 'test input' appears in field",
       lambda: (enter_text(driver, "test input"),
                assert_visible(driver, "test input")))

    tc("Verify UI", "TC-031", "EditText clears properly",
       "Field empty after clear()",
       lambda: (enter_text(driver, ""),
                _check_field_empty(driver)))

    tc("Verify UI", "TC-032", "Verify screen is vertically scrollable",
       "Swipe down completes without crash",
       lambda: swipe_down(driver))

    tc("Verify UI", "TC-033", "Verify screen scroll-up recovers",
       "Swipe up completes without crash",
       lambda: swipe_up(driver))

    tc("Verify UI", "TC-034", "Result card appears after first analysis",
       "'Analysis Result' visible after Analyze click",
       lambda: (analyze(driver, "Standard global news"),
                assert_visible(driver, "Analysis Result")))

    tc("Verify UI", "TC-035", "'Analysis Result' heading visible in card",
       "Card title text present",
       lambda: assert_visible(driver, "Analysis Result"))

    tc("Verify UI", "TC-036", "Result verdict element visible in card",
       "'Likely Genuine News' text present",
       lambda: assert_visible(driver, "Likely Genuine News"))

    tc("Verify UI", "TC-037", "Confidence score label visible in card",
       "'Confidence Score' text present",
       lambda: assert_visible(driver, "Confidence Score"))

    tc("Verify UI", "TC-038", "Recommendation label visible in card",
       "'Recommendation:' text present",
       lambda: assert_visible(driver, "Recommendation:"))

    tc("Verify UI", "TC-039", "Re-analysis updates result card",
       "New fake text replaces previous genuine result",
       lambda: (analyze(driver, "This is shocking fake news"),
                assert_visible(driver, "Likely Fake News")))

    go_back(driver)

    tc("Verify UI", "TC-040", "Back from Verify → Home screen clean",
       "TRUTHGUARD title visible after back",
       lambda: (click_btn(driver, "Verify News"),
                go_back(driver),
                assert_visible(driver, "TRUTHGUARD")))

    # ══════════════════════════════════════════════════════════════════════════
    # MODULE 3 · VERIFY NEWS — ANALYSIS LOGIC  (TC-041 – TC-075)
    # ══════════════════════════════════════════════════════════════════════════
    print("\n📋 MODULE 3: Verify News — Analysis Logic")
    click_btn(driver, "Verify News")

    # — Genuine result —
    tc("Verify Logic", "TC-041", "Genuine text → 'Likely Genuine News'",
       "No fake keywords → genuine result",
       lambda: (analyze(driver, "This is genuine news with no triggers"),
                assert_visible(driver, "Likely Genuine News")))

    tc("Verify Logic", "TC-042", "Genuine result: confidence 94%",
       "Confidence Score : 94% present",
       lambda: assert_visible(driver, "Confidence Score : 94%"))

    tc("Verify Logic", "TC-043", "Genuine result: '✅ Likely Genuine News' prefix",
       "Checkmark emoji in result text",
       lambda: assert_visible(driver, "✅ Likely Genuine News"))

    tc("Verify Logic", "TC-044", "Genuine result: 'This news appears reliable.'",
       "Recommendation text correct",
       lambda: assert_visible(driver, "This news appears reliable."))

    # — Fake keywords lowercase —
    tc("Verify Logic", "TC-045", "Keyword 'fake' (lowercase) → Fake result",
       "Text with 'fake' triggers Likely Fake News",
       lambda: (analyze(driver, "This article contains fake information"),
                assert_visible(driver, "Likely Fake News")))

    tc("Verify Logic", "TC-046", "Keyword 'hoax' (lowercase) → Fake result",
       "Text with 'hoax' triggers Likely Fake News",
       lambda: (analyze(driver, "This story is a hoax"),
                assert_visible(driver, "Likely Fake News")))

    tc("Verify Logic", "TC-047", "Keyword 'rumor' (lowercase) → Fake result",
       "Text with 'rumor' triggers Likely Fake News",
       lambda: (analyze(driver, "A rumor is spreading online"),
                assert_visible(driver, "Likely Fake News")))

    tc("Verify Logic", "TC-048", "Keyword 'clickbait' (lowercase) → Fake result",
       "Text with 'clickbait' triggers Likely Fake News",
       lambda: (analyze(driver, "Pure clickbait article"),
                assert_visible(driver, "Likely Fake News")))

    tc("Verify Logic", "TC-049", "Keyword 'shocking' (lowercase) → Fake result",
       "Text with 'shocking' triggers Likely Fake News",
       lambda: (analyze(driver, "A shocking revelation exposed today"),
                assert_visible(driver, "Likely Fake News")))

    # — Fake keywords UPPERCASE —
    tc("Verify Logic", "TC-050", "Keyword 'FAKE' (UPPERCASE) → Fake result",
       "Case-insensitive detection via .lowercase()",
       lambda: (analyze(driver, "This is completely FAKE information"),
                assert_visible(driver, "Likely Fake News")))

    tc("Verify Logic", "TC-051", "Keyword 'HOAX' (UPPERCASE) → Fake result",
       "Uppercase HOAX triggers Likely Fake News",
       lambda: (analyze(driver, "Experts say this is a HOAX"),
                assert_visible(driver, "Likely Fake News")))

    tc("Verify Logic", "TC-052", "Keyword 'RUMOR' (UPPERCASE) → Fake result",
       "Uppercase RUMOR triggers Likely Fake News",
       lambda: (analyze(driver, "RUMOR has it that prices will rise"),
                assert_visible(driver, "Likely Fake News")))

    tc("Verify Logic", "TC-053", "Keyword 'CLICKBAIT' (UPPERCASE) → Fake result",
       "Uppercase CLICKBAIT triggers Likely Fake News",
       lambda: (analyze(driver, "You won't believe this CLICKBAIT story"),
                assert_visible(driver, "Likely Fake News")))

    tc("Verify Logic", "TC-054", "Keyword 'SHOCKING' (UPPERCASE) → Fake result",
       "Uppercase SHOCKING triggers Likely Fake News",
       lambda: (analyze(driver, "SHOCKING news alert issued by media"),
                assert_visible(driver, "Likely Fake News")))

    # — Fake keywords Mixed-case —
    tc("Verify Logic", "TC-055", "Keyword 'fAkE' (mixed-case) → Fake result",
       "Mixed-case detected via .lowercase()",
       lambda: (analyze(driver, "News report: fAkE story spreading"),
                assert_visible(driver, "Likely Fake News")))

    tc("Verify Logic", "TC-056", "Keyword 'hOaX' (mixed-case) → Fake result",
       "Mixed-case hOaX triggers Likely Fake News",
       lambda: (analyze(driver, "Journalists confirm hOaX circulating"),
                assert_visible(driver, "Likely Fake News")))

    tc("Verify Logic", "TC-057", "Keyword 'RuMoR' (mixed-case) → Fake result",
       "Mixed-case RuMoR triggers Likely Fake News",
       lambda: (analyze(driver, "Online RuMoR sparks controversy"),
                assert_visible(driver, "Likely Fake News")))

    tc("Verify Logic", "TC-058", "Keyword 'ClIcKbAiT' (mixed-case) → Fake result",
       "Mixed-case ClIcKbAiT triggers Likely Fake News",
       lambda: (analyze(driver, "Editor labels article ClIcKbAiT"),
                assert_visible(driver, "Likely Fake News")))

    tc("Verify Logic", "TC-059", "Keyword 'ShOcKiNg' (mixed-case) → Fake result",
       "Mixed-case ShOcKiNg triggers Likely Fake News",
       lambda: (analyze(driver, "ShOcKiNg new alert from researchers"),
                assert_visible(driver, "Likely Fake News")))

    # — Fake result metadata —
    tc("Verify Logic", "TC-060", "Fake result: '❌ Likely Fake News' prefix",
       "Cross emoji in fake result text",
       lambda: (analyze(driver, "This is fake rumor spreading"),
                assert_visible(driver, "❌ Likely Fake News")))

    tc("Verify Logic", "TC-061", "Fake result: confidence 88%",
       "Confidence Score : 88% present",
       lambda: assert_visible(driver, "Confidence Score : 88%"))

    tc("Verify Logic", "TC-062", "Fake recommendation text correct",
       "Verify this news using trusted sources before sharing.",
       lambda: assert_visible(driver, "Verify this news using trusted sources before sharing."))

    tc("Verify Logic", "TC-063", "Recommendation label 'Recommendation:' present",
       "Label prefix present in card",
       lambda: assert_visible(driver, "Recommendation:"))

    # — Boundary conditions —
    tc("Verify Logic", "TC-064", "Empty input → Likely Genuine News",
       "No keywords in empty string → genuine",
       lambda: (analyze(driver, ""),
                assert_visible(driver, "Likely Genuine News")))

    tc("Verify Logic", "TC-065", "Whitespace-only → Likely Genuine News",
       "No keywords → genuine (spaces stripped by lowercase)",
       lambda: (analyze(driver, "   "),
                assert_visible(driver, "Likely Genuine News")))

    tc("Verify Logic", "TC-066", "Long genuine article → Genuine result",
       "200-char genuine text → Likely Genuine News",
       lambda: (analyze(driver,
                "Scientists at the international climate summit have announced new renewable "
                "energy targets, backed by strong economic data and verified research outcomes."),
                assert_visible(driver, "Likely Genuine News")))

    tc("Verify Logic", "TC-067", "Fake keyword at start of text → Fake result",
       "Text beginning with 'fake' triggers detection",
       lambda: (analyze(driver,
                "fake news story about climate summit with verified economic data"),
                assert_visible(driver, "Likely Fake News")))

    tc("Verify Logic", "TC-068", "Fake keyword at end of text → Fake result",
       "Text ending with 'hoax' still detected",
       lambda: (analyze(driver,
                "Breaking news about renewable energy policies — this is all a hoax"),
                assert_visible(driver, "Likely Fake News")))

    tc("Verify Logic", "TC-069", "Special characters around keyword → Fake",
       "### SHOCKING!!! $$$ triggers fake detection",
       lambda: (analyze(driver, "### SHOCKING!!! $$$ warning alert ***"),
                assert_visible(driver, "Likely Fake News")))

    tc("Verify Logic", "TC-070", "Multiple keywords in one text → Fake",
       "fake + hoax + rumor all present → Fake",
       lambda: (analyze(driver, "Warning fake hoax rumor clickbait shocking story"),
                assert_visible(driver, "Likely Fake News")))

    tc("Verify Logic", "TC-071", "Numeric-only text → Genuine",
       "Numbers have no keywords → genuine",
       lambda: (analyze(driver, "1234567890 9876543210"),
                assert_visible(driver, "Likely Genuine News")))

    tc("Verify Logic", "TC-072", "Single genuine word → Genuine",
       "Word 'science' → no keywords → genuine",
       lambda: (analyze(driver, "science"),
                assert_visible(driver, "Likely Genuine News")))

    tc("Verify Logic", "TC-073", "Single keyword 'fake' → Fake",
       "Minimal input still triggers detection",
       lambda: (analyze(driver, "fake"),
                assert_visible(driver, "Likely Fake News")))

    tc("Verify Logic", "TC-074", "Re-analysis Fake→Genuine updates result",
       "Switch from fake to genuine input → result changes",
       lambda: (analyze(driver, "total hoax"),
                assert_visible(driver, "Likely Fake News"),
                analyze(driver, "breaking science research news"),
                assert_visible(driver, "Likely Genuine News")))

    tc("Verify Logic", "TC-075", "Re-analysis Genuine→Fake updates result",
       "Switch from genuine to fake input → result changes",
       lambda: (analyze(driver, "genuine science report published"),
                assert_visible(driver, "Likely Genuine News"),
                analyze(driver, "clickbait rumor exposed"),
                assert_visible(driver, "Likely Fake News")))

    go_back(driver)

    # ══════════════════════════════════════════════════════════════════════════
    # MODULE 4 · TRENDING NEWS SCREEN  (TC-076 – TC-100)
    # ══════════════════════════════════════════════════════════════════════════
    print("\n📋 MODULE 4: Trending News Screen")
    click_btn(driver, "Trending News")

    tc("Trending", "TC-076", "Trending screen opens",
       "Trending News header visible",
       lambda: assert_visible(driver, "Trending News"))

    tc("Trending", "TC-077", "Header '📰 Trending News' exact text",
       "Emoji + title string correct",
       lambda: assert_visible(driver, "📰 Trending News"))

    tc("Trending", "TC-078", "BBC title: 'Scientists discover new climate monitoring technology'",
       "1st news card title exact match",
       lambda: assert_visible(driver, "Scientists discover new climate monitoring technology"))

    tc("Trending", "TC-079", "BBC source label: 'Source: BBC'",
       "1st card source label correct",
       lambda: assert_visible(driver, "Source: BBC"))

    tc("Trending", "TC-080", "Reuters title: 'AI transforming healthcare worldwide'",
       "2nd news card title exact match",
       lambda: assert_visible(driver, "AI transforming healthcare worldwide"))

    tc("Trending", "TC-081", "Reuters source label: 'Source: Reuters'",
       "2nd card source label correct",
       lambda: assert_visible(driver, "Source: Reuters"))

    tc("Trending", "TC-082", "NASA title: 'Space mission successfully reaches orbit'",
       "3rd news card title exact match",
       lambda: assert_visible(driver, "Space mission successfully reaches orbit"))

    tc("Trending", "TC-083", "NASA source label: 'Source: NASA'",
       "3rd card source label correct",
       lambda: assert_visible(driver, "Source: NASA"))

    tc("Trending", "TC-084", "Bloomberg title: 'Global economy shows positive growth'",
       "4th news card title exact match",
       lambda: assert_visible(driver, "Global economy shows positive growth"))

    tc("Trending", "TC-085", "Bloomberg source label: 'Source: Bloomberg'",
       "4th card source label correct",
       lambda: assert_visible(driver, "Source: Bloomberg"))

    tc("Trending", "TC-086", "UNESCO title: 'Education sector adopts AI learning tools'",
       "5th news card title exact match",
       lambda: assert_visible(driver, "Education sector adopts AI learning tools"))

    tc("Trending", "TC-087", "UNESCO source label: 'Source: UNESCO'",
       "5th card source label correct",
       lambda: assert_visible(driver, "Source: UNESCO"))

    tc("Trending", "TC-088", "All 5 source labels present on screen",
       "BBC, Reuters, NASA, Bloomberg, UNESCO all visible",
       lambda: [assert_visible(driver, f"Source: {s}")
                for s in ["BBC", "Reuters", "NASA", "Bloomberg", "UNESCO"]])

    tc("Trending", "TC-089", "Scroll-down gesture on Trending list",
       "Swipe down completes without crash",
       lambda: swipe_down(driver))

    tc("Trending", "TC-090", "Scroll-up gesture on Trending list",
       "Swipe up completes without crash",
       lambda: swipe_up(driver))

    tc("Trending", "TC-091", "Titles still visible after scroll cycle",
       "BBC card title accessible after scroll",
       lambda: assert_visible(driver, "Scientists discover new climate monitoring technology"))

    tc("Trending", "TC-092", "All 5 unique news titles present",
       "Each of the 5 card titles visible",
       lambda: [assert_visible(driver, t) for t in [
           "Scientists discover new climate monitoring technology",
           "AI transforming healthcare worldwide",
           "Space mission successfully reaches orbit",
           "Global economy shows positive growth",
           "Education sector adopts AI learning tools"]])

    tc("Trending", "TC-093", "No fake/hoax keywords appear in Trending content",
       "News items contain only legitimate content",
       lambda: (_assert_no_bad_word(driver, "hoax"),
                _assert_no_bad_word(driver, "clickbait"),
                _assert_no_bad_word(driver, "SHOCKING")))

    tc("Trending", "TC-094", "Trending loads within 5 seconds",
       "Source: BBC visible within 5s timeout",
       lambda: assert_visible(driver, "Source: BBC", timeout=5))

    tc("Trending", "TC-095", "Back from Trending → Home",
       "TRUTHGUARD title visible after back",
       lambda: (go_back(driver),
                assert_visible(driver, "TRUTHGUARD")))

    click_btn(driver, "Trending News")

    tc("Trending", "TC-096", "Trending screen re-entry works",
       "Re-navigating to Trending shows header",
       lambda: assert_visible(driver, "📰 Trending News"))

    tc("Trending", "TC-097", "Card content persists on second visit",
       "BBC title still present on re-entry",
       lambda: assert_visible(driver, "Scientists discover new climate monitoring technology"))

    tc("Trending", "TC-098", "All 5 source labels persist on re-entry",
       "No data loss on re-navigation",
       lambda: [assert_visible(driver, f"Source: {s}")
                for s in ["BBC", "Reuters", "NASA", "Bloomberg", "UNESCO"]])

    tc("Trending", "TC-099", "LazyColumn renders all 5 items",
       "All 5 cards present via lazy loading",
       lambda: [assert_visible(driver, t) for t in [
           "Scientists discover new climate monitoring technology",
           "AI transforming healthcare worldwide",
           "Space mission successfully reaches orbit",
           "Global economy shows positive growth",
           "Education sector adopts AI learning tools"]])

    go_back(driver)

    tc("Trending", "TC-100", "Trending → Home → Trending round-trip",
       "Full cycle: navigate out and back",
       lambda: (click_btn(driver, "Trending News"),
                assert_visible(driver, "📰 Trending News"),
                go_back(driver),
                assert_visible(driver, "TRUTHGUARD")))

    # ══════════════════════════════════════════════════════════════════════════
    # MODULE 5 · DASHBOARD SCREEN  (TC-101 – TC-115)
    # ══════════════════════════════════════════════════════════════════════════
    print("\n📋 MODULE 5: Dashboard Screen")
    click_btn(driver, "Dashboard")

    tc("Dashboard", "TC-101", "Dashboard screen opens",
       "Dashboard header visible",
       lambda: assert_visible(driver, "Dashboard"))

    tc("Dashboard", "TC-102", "Header '📊 Dashboard' exact text",
       "Emoji + title string correct",
       lambda: assert_visible(driver, "📊 Dashboard"))

    tc("Dashboard", "TC-103", "'Articles Verified' stat card title",
       "StatCard title text present",
       lambda: assert_visible(driver, "Articles Verified"))

    tc("Dashboard", "TC-104", "Articles Verified value = '25'",
       "Stat value 25 present",
       lambda: assert_visible(driver, "25"))

    tc("Dashboard", "TC-105", "'True News' stat card title",
       "StatCard title text present",
       lambda: assert_visible(driver, "True News"))

    tc("Dashboard", "TC-106", "True News value = '18'",
       "Stat value 18 present",
       lambda: assert_visible(driver, "18"))

    tc("Dashboard", "TC-107", "'Fake News' stat card title",
       "StatCard title text present",
       lambda: assert_visible(driver, "Fake News"))

    tc("Dashboard", "TC-108", "Fake News value = '7'",
       "Stat value 7 present",
       lambda: assert_visible(driver, "7"))

    tc("Dashboard", "TC-109", "'Accuracy' stat card title",
       "StatCard title text present",
       lambda: assert_visible(driver, "Accuracy"))

    tc("Dashboard", "TC-110", "Accuracy value = '92%'",
       "Stat value 92% present",
       lambda: assert_visible(driver, "92%"))

    tc("Dashboard", "TC-111", "All 4 stat card titles visible simultaneously",
       "Articles Verified, True News, Fake News, Accuracy all on screen",
       lambda: [assert_visible(driver, t)
                for t in ["Articles Verified", "True News", "Fake News", "Accuracy"]])

    tc("Dashboard", "TC-112", "All 4 stat values visible simultaneously",
       "25, 18, 7, 92% all on screen",
       lambda: [assert_visible(driver, v) for v in ["25", "18", "7", "92%"]])

    tc("Dashboard", "TC-113", "Dashboard persists on re-entry",
       "Navigate away and back — titles still visible",
       lambda: (go_back(driver),
                click_btn(driver, "Dashboard"),
                assert_visible(driver, "Articles Verified")))

    tc("Dashboard", "TC-114", "Dashboard stat values persist on re-entry",
       "Values 25, 18, 7, 92% still visible on second visit",
       lambda: [assert_visible(driver, v) for v in ["25", "18", "7", "92%"]])

    go_back(driver)

    tc("Dashboard", "TC-115", "Back from Dashboard → Home",
       "TRUTHGUARD title visible after back",
       lambda: (click_btn(driver, "Dashboard"),
                go_back(driver),
                assert_visible(driver, "TRUTHGUARD")))

    # ══════════════════════════════════════════════════════════════════════════
    # MODULE 6 · ABOUT SCREEN  (TC-116 – TC-126)
    # ══════════════════════════════════════════════════════════════════════════
    print("\n📋 MODULE 6: About Screen")
    click_btn(driver, "About")

    tc("About", "TC-116", "About screen opens",
       "TruthGuard title visible",
       lambda: assert_visible(driver, "TruthGuard"))

    tc("About", "TC-117", "Shield emoji '🛡️' visible on About",
       "Large emoji element present",
       lambda: assert_visible(driver, "🛡️"))

    tc("About", "TC-118", "'TruthGuard' title correct",
       "App name displayed prominently",
       lambda: assert_visible(driver, "TruthGuard"))

    tc("About", "TC-119", "'AI Powered Fake News Detection App' text",
       "Description text in card",
       lambda: assert_visible(driver, "AI Powered Fake News Detection App"))

    tc("About", "TC-120", "'Version : 1.0' text in card",
       "Version with space-colon-space format",
       lambda: assert_visible(driver, "Version : 1.0"))

    tc("About", "TC-121", "'Developed for Educational Purpose' text",
       "Purpose text in card",
       lambda: assert_visible(driver, "Developed for Educational Purpose"))

    tc("About", "TC-122", "'Technology : Kotlin + Jetpack Compose + AI' text",
       "Full technology stack text correct",
       lambda: assert_visible(driver, "Technology : Kotlin + Jetpack Compose + AI"))

    tc("About", "TC-123", "Copyright '© TruthGuard 2025' footer text",
       "Copyright symbol + year correct",
       lambda: assert_visible(driver, "© TruthGuard 2025"))

    tc("About", "TC-124", "All 4 About card items visible simultaneously",
       "All card texts on screen at once",
       lambda: [assert_visible(driver, t) for t in [
           "AI Powered Fake News Detection App",
           "Version : 1.0",
           "Developed for Educational Purpose",
           "Technology : Kotlin + Jetpack Compose + AI"]])

    tc("About", "TC-125", "About screen content persists on re-entry",
       "Navigate away and back — TruthGuard title still visible",
       lambda: (go_back(driver),
                click_btn(driver, "About"),
                assert_visible(driver, "TruthGuard")))

    go_back(driver)

    tc("About", "TC-126", "Back from About → Home",
       "TRUTHGUARD title visible after back",
       lambda: (click_btn(driver, "About"),
                go_back(driver),
                assert_visible(driver, "TRUTHGUARD")))

    # ══════════════════════════════════════════════════════════════════════════
    # MODULE 7 · DEVICE & ORIENTATION  (TC-127 – TC-130)
    # ══════════════════════════════════════════════════════════════════════════
    print("\n📋 MODULE 7: Device & Orientation")

    tc("Device", "TC-127", "Portrait→Landscape→Portrait: Home intact",
       "Rotate and restore — TRUTHGUARD still visible",
       lambda: (_rotate(driver, "LANDSCAPE"),
                assert_visible(driver, "TRUTHGUARD"),
                _rotate(driver, "PORTRAIT"),
                assert_visible(driver, "TRUTHGUARD")))

    tc("Device", "TC-128", "Landscape: Verify screen EditText accessible",
       "EditText present while in landscape",
       lambda: (driver.__setattr__("orientation", "LANDSCAPE"),
                time.sleep(1),
                click_btn(driver, "Verify News"),
                driver.find_element(AppiumBy.CLASS_NAME, "android.widget.EditText"),
                go_back(driver),
                driver.__setattr__("orientation", "PORTRAIT"),
                time.sleep(1)))

    tc("Device", "TC-129", "Home key press → app resumes",
       "KEYCODE_HOME then re-open app — TRUTHGUARD visible",
       lambda: (driver.press_keycode(3),
                time.sleep(1.5),
                driver.activate_app(APP_PACKAGE),
                time.sleep(1),
                assert_visible(driver, "TRUTHGUARD")))

    tc("Device", "TC-130", "Screen off/wake → app state intact",
       "KEYCODE_POWER off then on — app restored",
       lambda: (driver.press_keycode(26),
                time.sleep(1.5),
                driver.press_keycode(26),
                time.sleep(2),
                driver.activate_app(APP_PACKAGE),
                time.sleep(1),
                assert_visible(driver, "TRUTHGUARD")))

    # ══════════════════════════════════════════════════════════════════════════
    # MODULE 8 · END-TO-END USER FLOWS  (TC-131 – TC-135)
    # ══════════════════════════════════════════════════════════════════════════
    print("\n📋 MODULE 8: End-to-End User Flows")

    tc("E2E Flow", "TC-131", "Full fake-news flow: Home→Verify→Analyze→Back→Home",
       "Complete fake detection journey end-to-end",
       lambda: _e2e_fake(driver))

    tc("E2E Flow", "TC-132", "Full genuine-news flow: Home→Verify→Analyze→Back→Home",
       "Complete genuine detection journey end-to-end",
       lambda: _e2e_genuine(driver))

    tc("E2E Flow", "TC-133", "Sequential navigation of all 5 screens",
       "Visit every screen from Home and return",
       lambda: _e2e_all_screens(driver))

    tc("E2E Flow", "TC-134", "Verify screen state resets on re-entry",
       "Result card absent after navigating away and back",
       lambda: _e2e_state_reset(driver))

    tc("E2E Flow", "TC-135", "Dashboard→Home→Trending→Home round-trip",
       "Stats and news persist across navigation cycle",
       lambda: _e2e_dash_trending(driver))

    print("\n✅ All 135 test cases completed.")


# ─── Inline helpers (defined outside tc() to avoid closure/duplicate issues) ──
def _check_btn_count(driver):
    btns = driver.find_elements(AppiumBy.CLASS_NAME, "android.widget.Button")
    if len(btns) < 4:
        raise AssertionError(f"Expected >= 4 buttons, found {len(btns)}")


def _check_clickable(driver, text):
    el = find_text(driver, text)
    # Attempt to click the element; if it raises, consider it not clickable.
    try:
        el.click()
        # Return to previous screen to avoid side effects for subsequent navigation.
        try:
            go_back(driver)
        except Exception:
            pass
    except Exception as e:
        raise AssertionError(f"'{text}' is not clickable: {e}")


def _check_clickable_nodestructive(driver, text):
    """Check clickable/enabled attributes WITHOUT actually clicking — safe for
    buttons whose click has a side-effect that isn't a simple back-navigable screen."""
    el = find_text(driver, text)
    is_clickable = el.get_attribute("clickable")
    is_enabled = el.get_attribute("enabled")
    if is_clickable != "true" and is_enabled != "true":
        raise AssertionError(f"'{text}' is not clickable (clickable={is_clickable}, enabled={is_enabled})")


def _check_field_enabled(driver):
    # Wait for the EditText field to appear and be enabled
    el = wait_for_element(driver, AppiumBy.CLASS_NAME, "android.widget.EditText")
    if el.get_attribute("enabled") != "true":
        raise AssertionError("EditText is not enabled")


def _check_field_empty(driver):
    el = driver.find_element(AppiumBy.CLASS_NAME, "android.widget.EditText")
    val = el.get_attribute("text") or ""
    if val.strip():
        raise AssertionError(f"EditText not empty, contains: '{val}'")


def _nav_to(driver, btn_text, expected_header):
    click_btn(driver, btn_text)
    assert_visible(driver, expected_header)


def _assert_no_bad_word(driver, word):
    try:
        find_text(driver, word, timeout=2)
        raise AssertionError(f"Unexpected keyword '{word}' found in Trending content")
    except AssertionError as e:
        if "Unexpected" in str(e):
            raise


def _rotate(driver, orientation):
    driver.orientation = orientation
    time.sleep(1.5)


def _ensure_home(driver):
    driver.activate_app(APP_PACKAGE)
    for _ in range(3):
        if driver.find_elements(AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().text("TRUTHGUARD")'):
            return
        try:
            driver.back()
            time.sleep(0.5)
        except Exception:
            pass


def _e2e_fake(driver):
    _ensure_home(driver)
    click_btn(driver, "Verify News")
    analyze(driver, "Breaking: hoax about global leaders proven false")
    assert_visible(driver, "❌ Likely Fake News")
    assert_visible(driver, "Confidence Score : 88%")
    assert_visible(driver, "Verify this news using trusted sources before sharing.")
    go_back(driver)
    assert_visible(driver, "TRUTHGUARD")


def _e2e_genuine(driver):
    _ensure_home(driver)
    click_btn(driver, "Verify News")
    analyze(driver, "NASA confirms successful satellite launch for climate monitoring")
    assert_visible(driver, "✅ Likely Genuine News")
    assert_visible(driver, "Confidence Score : 94%")
    assert_visible(driver, "This news appears reliable.")
    go_back(driver)
    assert_visible(driver, "TRUTHGUARD")


def _e2e_all_screens(driver):
    _ensure_home(driver)
    pairs = [
        ("Verify News",  "Verify News"),
        ("Trending News", "📰 Trending News"),
        ("Dashboard",    "📊 Dashboard"),
        ("About",        "TruthGuard"),
    ]
    for btn, header in pairs:
        click_btn(driver, btn)
        assert_visible(driver, header)
        go_back(driver)
        assert_visible(driver, "TRUTHGUARD")


def _e2e_state_reset(driver):
    _ensure_home(driver)
    click_btn(driver, "Verify News")
    analyze(driver, "shocking rumor exposed by investigator")
    assert_visible(driver, "❌ Likely Fake News")
    go_back(driver)
    click_btn(driver, "Verify News")
    assert_visible(driver, "Paste News Here")   # field label resets
    # result card should be gone (Compose remember {} resets on recomposition)
    try:
        find_text(driver, "Analysis Result", timeout=2)
        raise AssertionError("State not reset — Analysis Result still visible on re-entry")
    except AssertionError as e:
        if "State not reset" in str(e):
            raise
    go_back(driver)


def _e2e_dash_trending(driver):
    _ensure_home(driver)
    click_btn(driver, "Dashboard")
    assert_visible(driver, "92%")
    go_back(driver)
    click_btn(driver, "Trending News")
    assert_visible(driver, "Source: BBC")
    go_back(driver)
    assert_visible(driver, "TRUTHGUARD")


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL REPORT GENERATOR
# ═══════════════════════════════════════════════════════════════════════════════
def generate_excel_report(report_path):
    print("\n📊 Generating Excel report...")
    wb = openpyxl.Workbook()

    def fnt(name="Plus Jakarta Sans", size=10, bold=False, color="000000"):
        return Font(name=name, size=size, bold=bold, color=color)

    def fill(hex_color):
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    def bdr():
        s = Side(style='thin', color=COLORS['border'])
        return Border(left=s, right=s, top=s, bottom=s)

    def aln(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    passed_cnt   = len([r for r in results if r['status'] == 'PASSED'])
    failed_cnt   = len([r for r in results if r['status'] == 'FAILED'])
    total_cnt    = len(results)
    pass_rate    = round((passed_cnt / total_cnt) * 100) if total_cnt > 0 else 0
    duration_sec = round(time.time() - suite_start, 1)

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 1 — SUMMARY
    # ─────────────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "📊 Summary"
    ws.views.sheetView[0].showGridLines = False

    # Banner
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = "🛡️  TRUTHGUARD ANDROID — APPIUM E2E TEST REPORT"
    c.font      = fnt("Outfit", 18, True, COLORS['headerText'])
    c.fill      = fill(COLORS['headerBg'])
    c.alignment = aln("center")
    ws.row_dimensions[1].height = 52

    # Sub-banner
    ws.merge_cells("A2:H2")
    ts  = datetime.now().strftime("%d/%m/%Y, %I:%M:%S %p")
    sub = ws["A2"]
    sub.value     = (f"Generated: {ts}  |  Engine: Python + Appium Client 3.x  |  "
                     f"Platform: Android API 33  |  Automation: UIAutomator2")
    sub.font      = fnt("Plus Jakarta Sans", 10, color="B0C4DE")
    sub.fill      = fill(COLORS['subHeaderBg'])
    sub.alignment = aln("center")
    ws.row_dimensions[2].height = 26

    ws.append([])  # row 3 spacer

    # KPI boxes
    kpis = [
        ("TOTAL TESTS", total_cnt,       COLORS['statTotal'], "B"),
        ("PASSED",      passed_cnt,      COLORS['statPass'],  "D"),
        ("FAILED",      failed_cnt,      COLORS['statFail'],  "F"),
        ("PASS RATE",   f"{pass_rate}%", COLORS['accent'],    "H"),
    ]
    for label, val, color, col in kpis:
        ws.merge_cells(f"{col}4:{col}5")
        kc = ws[f"{col}4"]
        kc.value     = val
        kc.font      = fnt("Outfit", 28, True, "FFFFFF")
        kc.fill      = fill(color)
        kc.alignment = aln("center")
        lc = ws[f"{col}6"]
        lc.value     = label
        lc.font      = fnt("Plus Jakarta Sans", 9, True, color)
        lc.alignment = aln("center")
    ws.row_dimensions[4].height = 44
    ws.row_dimensions[5].height = 44
    ws.row_dimensions[6].height = 18

    ws.append([])
    ws.append([])

    # Execution details
    dr = ws.max_row + 1
    ws.append(["", "EXECUTION DETAILS", "", "", "", "", "", ""])
    ws.merge_cells(f"B{dr}:H{dr}")
    ws[f"B{dr}"].font  = fnt("Plus Jakarta Sans", 12, True, COLORS['accent'])
    ws[f"B{dr}"].fill  = fill("E8EEF7")
    ws.row_dimensions[dr].height = 26

    details = [
        ("Test Suite",       "TruthGuard Android — Appium E2E"),
        ("App Package",      APP_PACKAGE),
        ("Total Tests",      total_cnt),
        ("Passed",           passed_cnt),
        ("Failed",           failed_cnt),
        ("Pass Rate",        f"{pass_rate}%"),
        ("Duration",         f"{duration_sec}s"),
        ("Automation",       "UIAutomator2 / Appium 2.x"),
        ("Android API",      "33"),
        ("Test Language",    "Python 3.12"),
        ("Report Generated", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
    ]
    for lbl, val in details:
        ws.append(["", lbl, "", val])
        r = ws.max_row
        ws.row_dimensions[r].height = 20
        ws[f"B{r}"].font  = fnt(bold=True)
        ws[f"B{r}"].fill  = fill("F0F4FF")
        ws[f"D{r}"].font  = fnt()
        ws.merge_cells(f"D{r}:H{r}")

    ws.append([])

    # Module breakdown
    mb = ws.max_row + 1
    ws.append(["", "MODULE BREAKDOWN", "", "", "", "", "", ""])
    ws.merge_cells(f"B{mb}:H{mb}")
    ws[f"B{mb}"].font  = fnt("Plus Jakarta Sans", 12, True, COLORS['accent'])
    ws[f"B{mb}"].fill  = fill("E8EEF7")
    ws.row_dimensions[mb].height = 26

    ws.append(["", "Module", "", "Total", "Passed", "Failed", "Pass Rate", ""])
    hr = ws.max_row
    ws.row_dimensions[hr].height = 22
    for col in ["B", "D", "E", "F", "G"]:
        c = ws[f"{col}{hr}"]
        c.font      = fnt("Plus Jakarta Sans", 10, True, "FFFFFF")
        c.fill      = fill(COLORS['accent'])
        c.alignment = aln("center")

    modules = list(dict.fromkeys([r['module'] for r in results]))
    for mod in modules:
        mc  = [r for r in results if r['module'] == mod]
        mp  = len([r for r in mc if r['status'] == 'PASSED'])
        mf  = len([r for r in mc if r['status'] == 'FAILED'])
        mr  = f"{round((mp / len(mc)) * 100)}%"
        ws.append(["", mod, "", len(mc), mp, mf, mr, ""])
        rn = ws.max_row
        ws.row_dimensions[rn].height = 20
        ws[f"B{rn}"].font = fnt(bold=True)
        for col in ["D", "E", "F", "G"]:
            ws[f"{col}{rn}"].alignment = aln("center")
        ws[f"E{rn}"].font = fnt(color=COLORS['statPass'])
        ws[f"F{rn}"].font = fnt(color=COLORS['statFail'])
        ws[f"G{rn}"].font = fnt(bold=True)

    for col, width in zip("ABCDEFGH", [2, 34, 4, 18, 14, 14, 14, 2]):
        ws.column_dimensions[col].width = width

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 2 — ALL TEST CASES
    # ─────────────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("🧪 Test Cases")
    ws2.views.sheetView[0].showGridLines = True

    ws2.merge_cells("A1:H1")
    ws2["A1"].value     = "🛡️  TRUTHGUARD ANDROID — ALL TEST CASES DETAIL"
    ws2["A1"].font      = fnt("Outfit", 15, True, COLORS['headerText'])
    ws2["A1"].fill      = fill(COLORS['headerBg'])
    ws2["A1"].alignment = aln("center")
    ws2.row_dimensions[1].height = 44

    ws2.merge_cells("A2:H2")
    info = ws2["A2"]
    info.value     = (f"Total: {total_cnt}  |  Passed: {passed_cnt}  |  "
                      f"Failed: {failed_cnt}  |  Pass Rate: {pass_rate}%  |  Duration: {duration_sec}s")
    info.font      = fnt(size=10, color="FFFFFF")
    info.fill      = fill(COLORS['subHeaderBg'])
    info.alignment = aln("center")
    ws2.row_dimensions[2].height = 22

    headers = ["#", "Test ID", "Module", "Test Case Name",
               "Description", "Status", "Duration (ms)", "Error / Notes"]
    ws2.append(headers)
    hr2 = ws2.max_row
    ws2.row_dimensions[hr2].height = 30
    for cell in ws2[hr2]:
        cell.font      = fnt("Plus Jakarta Sans", 11, True, "FFFFFF")
        cell.fill      = fill(COLORS['accent'])
        cell.alignment = aln("center")
        cell.border    = Border(bottom=Side(style='medium', color=COLORS['moduleText']))

    row_idx  = 0
    cur_mod  = None
    for r in results:
        row_idx += 1
        if r['module'] != cur_mod:
            cur_mod = r['module']
            ws2.append(["", "", cur_mod.upper(), "", "", "", "", ""])
            sn = ws2.max_row
            ws2.merge_cells(f"C{sn}:H{sn}")
            ws2.row_dimensions[sn].height = 24
            for col in range(1, 9):
                c2 = ws2.cell(sn, col)
                c2.fill = fill(COLORS['moduleBg'])
                c2.font = fnt("Plus Jakarta Sans", 10, True, COLORS['moduleText'])

        is_pass = r['status'] == 'PASSED'
        s_bg    = COLORS['passedBg']   if is_pass else COLORS['failedBg']
        s_fg    = COLORS['passedText'] if is_pass else COLORS['failedText']
        r_bg    = COLORS['rowAlt'] if row_idx % 2 == 0 else COLORS['rowNorm']

        ws2.append([row_idx, r['id'], r['module'], r['name'],
                    r['desc'], r['status'], r['duration'], r['error']])
        dn = ws2.max_row
        ws2.row_dimensions[dn].height = 22

        for col in range(1, 9):
            c2 = ws2.cell(dn, col)
            c2.border = bdr()
            if col == 6:
                c2.font      = fnt(bold=True, color=s_fg)
                c2.fill      = fill(s_bg)
                c2.alignment = aln("center")
            elif col == 8 and r['error']:
                c2.font      = fnt(size=9, color=COLORS['failedText'])
                c2.fill      = fill(r_bg)
                c2.alignment = aln(wrap=True)
            else:
                c2.font  = fnt()
                c2.fill  = fill(r_bg)
                if col in (1, 2):
                    c2.alignment = aln("center")
                elif col == 7:
                    c2.alignment = aln("right")

    for col, width in zip("ABCDEFGH", [5, 10, 16, 42, 52, 14, 14, 55]):
        ws2.column_dimensions[col].width = width

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 3 — FAILED TESTS (only when failures exist)
    # ─────────────────────────────────────────────────────────────────────────
    failed_list = [r for r in results if r['status'] == 'FAILED']
    if failed_list:
        ws3 = wb.create_sheet("❌ Failed Tests")
        ws3.merge_cells("A1:G1")
        ws3["A1"].value     = f"❌  FAILED TEST CASES — {len(failed_list)} Failures"
        ws3["A1"].font      = fnt("Outfit", 14, True, "FFFFFF")
        ws3["A1"].fill      = fill(COLORS['statFail'])
        ws3["A1"].alignment = aln("center")
        ws3.row_dimensions[1].height = 40

        ws3.append(["#", "Test ID", "Module", "Test Case Name",
                    "Status", "Duration (ms)", "Error Message"])
        fh = ws3.max_row
        ws3.row_dimensions[fh].height = 26
        for cell in ws3[fh]:
            cell.font      = fnt("Plus Jakarta Sans", 10, True, "FFFFFF")
            cell.fill      = fill(COLORS['accent'])
            cell.alignment = aln("center")

        for i, r in enumerate(failed_list):
            ws3.append([i + 1, r['id'], r['module'], r['name'],
                        r['status'], r['duration'], r['error']])
            fn = ws3.max_row
            ws3.row_dimensions[fn].height = 30
            for col in range(1, 8):
                c3 = ws3.cell(fn, col)
                c3.fill   = fill(COLORS['failedBg'])
                c3.border = bdr()
                if col == 5:
                    c3.font      = fnt(bold=True, color=COLORS['failedText'])
                    c3.alignment = aln("center")
                elif col == 7:
                    c3.font      = fnt(size=9, color=COLORS['failedText'])
                    c3.alignment = aln(wrap=True)
                else:
                    c3.font = fnt()

        for col, width in zip("ABCDEFG", [5, 10, 16, 44, 12, 14, 62]):
            ws3.column_dimensions[col].width = width

    wb.save(report_path)
    print(f"📄 Excel report saved → {report_path}")


# ═══════════════════════════════════════════════════════════════════════════════
#  MARKDOWN SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════
def generate_markdown_summary(summary_path):
    print("\n📝 Generating Markdown summary...")
    passed    = len([r for r in results if r['status'] == 'PASSED'])
    failed    = len([r for r in results if r['status'] == 'FAILED'])
    total     = len(results)
    pass_rate = round((passed / total) * 100) if total > 0 else 0
    duration  = round(time.time() - suite_start, 1)
    timestamp = datetime.now().strftime("%d/%m/%Y, %I:%M:%S %p")
    badge     = '🟢' if pass_rate == 100 else '🟡' if pass_rate >= 80 else '🔴'

    md  = "# 🛡️ TruthGuard Android — Appium E2E Test Report\n\n"
    md += (f"> **Generated:** {timestamp} &nbsp;|&nbsp; "
           f"**Platform:** Android API 33 &nbsp;|&nbsp; "
           f"**Engine:** Python + Appium Client 3.x\n\n")
    md += "---\n\n"

    md += "## 📊 Results Summary\n\n"
    md += f"| {badge} Pass Rate | 📋 Total Tests | ✅ Passed | ❌ Failed | ⏱️ Duration |\n"
    md += "|:-----------:|:--------------:|:---------:|:---------:|:----------:|\n"
    md += f"| **{pass_rate}%** | **{total}** | **{passed}** | **{failed}** | **{duration}s** |\n\n"

    md += "## 📋 Module Breakdown\n\n"
    md += "| Module | Tests | ✅ Passed | ❌ Failed | Pass Rate |\n"
    md += "|--------|:-----:|:---------:|:---------:|:---------:|\n"
    modules = list(dict.fromkeys([r['module'] for r in results]))
    for mod in modules:
        mc   = [r for r in results if r['module'] == mod]
        mp   = len([r for r in mc if r['status'] == 'PASSED'])
        mf   = len([r for r in mc if r['status'] == 'FAILED'])
        rate = round((mp / len(mc)) * 100)
        icon = '✅' if mf == 0 else '❌'
        md += f"| {icon} {mod} | {len(mc)} | {mp} | {mf} | {rate}% |\n"
    md += "\n"

    failed_list = [r for r in results if r['status'] == 'FAILED']
    if failed_list:
        md += "## ❌ Failed Test Cases\n\n"
        md += "| Test ID | Module | Test Name | Error |\n"
        md += "|---------|--------|-----------|-------|\n"
        for r in failed_list:
            err = (r['error'] or '').replace("|", "\\|")[:120]
            md += f"| `{r['id']}` | {r['module']} | {r['name']} | `{err}` |\n"
        md += "\n"
    else:
        md += "## 🎉 All Tests Passed!\n\n"
        md += (f"> All **{total}** E2E test cases passed with a **{pass_rate}%** "
               f"pass rate on the TruthGuard Android app.\n\n")

    md += "---\n"
    md += ("*Excel report available as a downloadable run artifact — "
           "see the **Artifacts** section below this run.*\n")

    with open(summary_path, "w", encoding="utf-8") as f:
        f.write(md)
    print(f"📝 Markdown summary saved → {summary_path}")


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    print("==========================================================")
    print("   🛡️  TRUTHGUARD ANDROID — APPIUM E2E TEST RUNNER")
    print("   135 Test Cases  |  8 Modules  |  Python 3.12 + Appium")
    print("==========================================================")
    print(f"ℹ️  CI mode  : {IS_CI}")
    print(f"ℹ️  ADB      : {ADB_PATH}")
    print(f"ℹ️  APK      : {APK_PATH}")
    print(f"ℹ️  AVD Name : {AVD_NAME}")
    print(f"ℹ️  Package  : {APP_PACKAGE}")

    try:
        start_emulator()
    except Exception as e:
        print(f"🔴 Emulator error: {e} — continuing")

    server_proc, server_log = None, None
    try:
        server_proc, server_log = start_appium_server()
    except Exception as e:
        print(f"🔴 Appium error: {e} — continuing")

    driver = None
    try:
        print("\n🧪 Connecting to UIAutomator2 driver...")
        opts = UiAutomator2Options()
        opts.platform_name   = "Android"
        opts.device_name     = AVD_NAME
        opts.app             = APK_PATH
        opts.automation_name = "UiAutomator2"
        opts.set_capability("autoGrantPermissions",             True)
        opts.set_capability("uiautomator2ServerLaunchTimeout",  90000)
        opts.set_capability("uiautomator2ServerInstallTimeout", 90000)
        opts.set_capability("adbExecTimeout",                   60000)
        opts.set_capability("newCommandTimeout",                300)

        driver = webdriver.Remote(
            f"http://{APPIUM_HOST}:{APPIUM_PORT}", options=opts)
        print("🟢 Driver connected.")
        time.sleep(2)

        run_all_tests(driver)

    except Exception as e:
        print(f"\n🔴 Critical error: {e}")
    finally:
        if driver:
            print("\n🧹 Quitting driver...")
            try:
                driver.quit()
            except Exception:
                pass

        if server_proc:
            print("🧹 Stopping Appium server...")
            server_proc.terminate()
        if server_log:
            server_log.close()

        passed    = len([r for r in results if r['status'] == 'PASSED'])
        failed    = len([r for r in results if r['status'] == 'FAILED'])
        total     = len(results)
        duration  = round(time.time() - suite_start, 1)
        pass_rate = round((passed / total) * 100) if total > 0 else 0

        print("\n==========================================================")
        print(f"  📊 RESULTS  |  Total: {total}  ✅ Passed: {passed}  ❌ Failed: {failed}")
        print(f"  ⏱️  Duration: {duration}s  |  Pass Rate: {pass_rate}%")
        print("==========================================================")

        ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_dir    = os.path.dirname(os.path.abspath(__file__))
        report_xlsx = os.path.join(base_dir, f"Appium_E2E_Report_TruthGuard_{ts}.xlsx")
        report_md   = os.path.join(base_dir, "test-summary.md")

        generate_excel_report(report_xlsx)
        generate_markdown_summary(report_md)

        print("\n✨ Testing complete! Open the .xlsx report for full details.")

        if failed > 0 or total == 0:
            print("\n❌ Suite failed: one or more tests failed or no tests were run.", flush=True)
            sys.exit(1)


if __name__ == "__main__":
    main()
